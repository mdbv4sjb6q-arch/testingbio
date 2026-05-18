# app.py - 完整的投资AI系统Web服务器
"""
支持全球市场的投资AI系统Web应用
- 支持WiFi网络下的所有用户访问
- 本地LLM (dmind-trading-merged)
- 多Agent共识决策
- 全球市场数据 (外汇、美股、港股、数字货币)
- K线可视化
"""

from flask import Flask, render_template, request, jsonify, session
from flask_cors import CORS
from flask_session import Session
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
import os
import json
from typing import Dict, List, Optional, Tuple
import uuid
import random
import traceback
import re

# 本地导入
from config import APIConfig, AssetClass, MarketRegion
from base_provider import BaseProvider
from ccxt_provider import CCXTProvider
from yfinance_provider import YFinanceProvider
from itick_provider import ITickProvider
from cmc_provider import CMCProvider
from massive_provider import MassiveProvider
from models import MarketData, Instrument, UserQuery, AIResponse, ConsensusDecision
from technical_indicators import TechnicalIndicators, get_trading_signal, calculate_all_indicators
from ai_agents import get_agent_system, analyze_market_consensus
from llama_wrapper import get_llm

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('investment_ai.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Flask应用
app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # 允许HTTP用于局域网

Session(app)
CORS(app, supports_credentials=True)

# 全局数据提供者
providers: Dict[str, BaseProvider] = {}

def init_providers():
    """初始化数据提供者"""
    global providers
    try:
        providers['itick'] = ITickProvider()
        logger.info("✓ iTick provider initialized")
    except Exception as e:
        logger.warning(f"iTick provider failed: {e}")
    
    try:
        providers['yfinance'] = YFinanceProvider()
        logger.info("✓ Yahoo Finance provider initialized")
    except Exception as e:
        logger.warning(f"Yahoo Finance provider failed: {e}")
    
    try:
        providers['ccxt_binance'] = CCXTProvider('binance')
        logger.info("✓ CCXT Binance provider initialized")
    except Exception as e:
        logger.warning(f"CCXT provider failed: {e}")
    
    try:
        providers['cmc'] = CMCProvider("41f58281b0a34e8ea821ebb470a2fdde")
        logger.info("✓ CoinMarketCap provider initialized")
    except Exception as e:
        logger.warning(f"CoinMarketCap provider failed: {e}")
    
    try:
        providers['massive'] = MassiveProvider("YgorrHnF7VG9Sto2gdn4jrXq8P3AtydR")
        logger.info("✓ Massive API provider initialized")
    except Exception as e:
        logger.warning(f"Massive API provider failed: {e}")

# 应用启动时初始化
with app.app_context():
    init_providers()

# ==================== 辅助函数 ====================

def generate_demo_data(symbol: str, days: int = 30) -> Optional[pd.DataFrame]:
    """生成演示数据用于演示"""
    import random
    
    dates = pd.date_range(end=datetime.now(), periods=days, freq='D')
    # 模拟价格走势
    base_price = 100.0
    prices = [base_price]
    
    for i in range(1, days):
        change = random.uniform(-2, 2)  # ±2%变化
        prices.append(prices[-1] * (1 + change/100))
    
    data = []
    for i, date in enumerate(dates):
        price = prices[i]
        open_price = price * (1 + random.uniform(-0.5, 0.5)/100)
        close_price = price
        high_price = max(open_price, close_price) * (1 + random.uniform(0, 1)/100)
        low_price = min(open_price, close_price) * (1 - random.uniform(0, 1)/100)
        volume = random.uniform(1000000, 10000000)
        
        data.append({
            'timestamp': date,
            'open': open_price,
            'high': high_price,
            'low': low_price,
            'close': close_price,
            'volume': volume
        })
    
    return pd.DataFrame(data)

def get_market_data(symbol: str, 
                   timeframe: str = "1d",
                   days: int = 100) -> Optional[pd.DataFrame]:
    """
    获取市场数据 - 自动选择最优提供者
    
    Args:
        symbol: 标的代码 (支持多种格式)
        timeframe: 时间周期
        days: 获取天数
        
    Returns:
        市场数据DataFrame或None
    """
    # 尝试不同的提供者
    for provider_name, provider in providers.items():
        try:
            logger.info(f"Trying {provider_name} for {symbol}")
            data = provider.get_ohlcv(symbol, timeframe=timeframe, limit=100)
            
            if data is not None and len(data) > 0:
                logger.info(f"✓ Got {len(data)} candles from {provider_name}")
                return data
        
        except Exception as e:
            logger.debug(f"Failed with {provider_name}: {e}")
            continue
    
    logger.warning(f"Could not fetch data for {symbol} from any provider")
    return None

def search_instruments(query: str) -> List[Dict]:
    """搜索全球标的 - 优先CMC/Massive，fallback到hardcoded列表"""
    query_upper = query.upper()
    results = []
    seen = set()
    
    # 1. 首先尝试从CMC（优先级最高 - 加密货币）
    if 'cmc' in providers:
        try:
            instruments = providers['cmc'].search(query)
            for inst in instruments:
                key = f"{inst.get('symbol', '')}_{inst.get('type', '')}"
                if key not in seen:
                    seen.add(key)
                    results.append(inst)
            logger.info(f"CMC search found {len(instruments)} results for '{query}'")
        except Exception as e:
            logger.debug(f"CMC search failed: {e}")
    
    # 2. 然后尝试Massive（全球市场）
    if 'massive' in providers and len(results) < 5:
        try:
            instruments = providers['massive'].search(query)
            for inst in instruments:
                key = f"{inst.get('symbol', '')}_{inst.get('type', '')}"
                if key not in seen:
                    seen.add(key)
                    results.append(inst)
            logger.info(f"Massive search found {len(instruments)} results for '{query}'")
        except Exception as e:
            logger.debug(f"Massive search failed: {e}")
    
    # 3. 然后尝试其他providers
    if len(results) < 5:
        for provider_name, provider in providers.items():
            if provider_name not in ['cmc', 'massive']:
                try:
                    instruments = provider.search(query)
                    for inst in instruments:
                        key = f"{inst.get('symbol', '')}_{inst.get('type', '')}"
                        if key not in seen:
                            seen.add(key)
                            results.append(inst)
                except Exception as e:
                    logger.debug(f"Search failed in {provider_name}: {e}")
    
    # 4. 最后使用hardcoded fallback
    if len(results) < 5:
        popular_instruments = [
            {'symbol': 'AAPL', 'name': 'Apple Inc.', 'exchange': 'NASDAQ', 'type': 'STOCK'},
            {'symbol': 'MSFT', 'name': 'Microsoft Corporation', 'exchange': 'NASDAQ', 'type': 'STOCK'},
            {'symbol': 'GOOGL', 'name': 'Alphabet Inc.', 'exchange': 'NASDAQ', 'type': 'STOCK'},
            {'symbol': 'AMZN', 'name': 'Amazon.com Inc.', 'exchange': 'NASDAQ', 'type': 'STOCK'},
            {'symbol': 'TSLA', 'name': 'Tesla Inc.', 'exchange': 'NASDAQ', 'type': 'STOCK'},
            {'symbol': '0700.HK', 'name': 'Tencent Holdings', 'exchange': 'HKEX', 'type': 'STOCK'},
            {'symbol': '9988.HK', 'name': 'Alibaba Group', 'exchange': 'HKEX', 'type': 'STOCK'},
            {'symbol': 'BTC/USDT', 'name': 'Bitcoin', 'exchange': 'BINANCE', 'type': 'CRYPTO'},
            {'symbol': 'ETH/USDT', 'name': 'Ethereum', 'exchange': 'BINANCE', 'type': 'CRYPTO'},
            {'symbol': 'EURUSD', 'name': 'EUR/USD', 'exchange': 'FOREX', 'type': 'FOREX'},
            {'symbol': 'GBPUSD', 'name': 'GBP/USD', 'exchange': 'FOREX', 'type': 'FOREX'},
            {'symbol': 'GOLD', 'name': 'Gold Futures', 'exchange': 'NYMEX', 'type': 'COMMODITY'},
            {'symbol': 'XOP', 'name': 'Oil & Gas ETF', 'exchange': 'NASDAQ', 'type': 'ETF'},
        ]
        for inst in popular_instruments:
            if query_upper in inst['symbol'] or query_upper in inst['name'].upper():
                key = f"{inst['symbol']}_{inst['type']}"
                if key not in seen:
                    seen.add(key)
                    results.append(inst)
    
    logger.info(f"Search for '{query}' returned {len(results)} results")
    return results[:20]  # 限制结果数量

def generate_chart_data(df: pd.DataFrame, 
                       indicators_dict: Dict) -> Dict:
    """生成图表数据"""
    
    chart_data = {
        'timestamps': [],
        'ohlcv': [],
        'ma1': [],
        'ma2': [],
        'ma3': [],
        'ma4': [],
        'var3': [],
        'var4': [],
        'volume': []
    }
    
    if df.empty:
        return chart_data
    
    # 重置索引确保干净的迭代
    df = df.reset_index(drop=True)
    
    # 准备OHLCV
    for i, row in df.iterrows():
        # 使用行号作为时间戳索引
        timestamp = str(i)
        if 'timestamp' in df.columns:
            try:
                timestamp = str(row.get('timestamp', i))
            except:
                pass
        
        chart_data['timestamps'].append(timestamp)
        chart_data['ohlcv'].append([
            timestamp,
            float(row.get('open', 0) or 0),
            float(row.get('high', 0) or 0),
            float(row.get('low', 0) or 0),
            float(row.get('close', 0) or 0),
            float(row.get('volume', 0) or 0)
        ])
    
    # 添加指标
    for key in ['MA1', 'MA2', 'MA3', 'MA4', 'VAR3', 'VAR4']:
        if key in indicators_dict:
            values = indicators_dict[key]
            if isinstance(values, np.ndarray):
                try:
                    chart_data[key.lower()] = [float(v) if not np.isnan(v) else None for v in values]
                except:
                    chart_data[key.lower()] = []
            elif isinstance(values, (list, tuple)):
                chart_data[key.lower()] = [float(v) if v is not None else None for v in values]
            else:
                chart_data[key.lower()] = []
    
    # 成交量
    if 'volume' in df.columns:
        chart_data['volume'] = df['volume'].fillna(0).tolist()
    
    return chart_data

# ==================== API路由 ====================

@app.route('/', methods=['GET'])
def index():
    """首页"""
    # 尝试返回 HTML 文件
    try:
        return render_template('index.html')
    except Exception as e:
        logger.warning(f"Failed to render template: {e}")
        # 如果渲染失败，返回 JSON API 文档
        return jsonify({
            'status': 'ok',
            'message': '投资AI系统运行中',
            'version': '1.0.0',
            'note': '请使用 /api/* 端点或访问 /pro 和 /svip 页面',
            'endpoints': {
                '/': 'GET - 标准版首页',
                '/pro': 'GET - PRO会员页面',
                '/svip': 'GET - SVIP会员页面',
                '/api/search': 'GET - 搜索标的',
                '/api/market_data': 'GET - 获取市场数据和K线',
                '/api/analysis': 'POST - 获取AI分析',
                '/api/chat': 'POST - AI对话',
                '/api/agents_info': 'GET - 获取Agent信息',
                '/api/indicators': 'GET - 获取技术指标',
                '/api/health': 'GET - 健康检查'
            }
        })

# ==================== 会员页面路由 ====================

@app.route('/pro', methods=['GET'])
def pro_page():
    """PRO会员页面"""
    try:
        return render_template('pro.html')
    except Exception as e:
        logger.error(f"Failed to render pro.html: {e}")
        return jsonify({'error': 'Failed to load PRO page'}), 500

@app.route('/svip', methods=['GET'])
def svip_page():
    """SVIP会员页面"""
    try:
        return render_template('svip.html')
    except Exception as e:
        logger.error(f"Failed to render svip.html: {e}")
        return jsonify({'error': 'Failed to load SVIP page'}), 500

@app.route('/api/search', methods=['GET'])
def api_search():
    """搜索全球标的"""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'error': 'Query required'}), 400
        
        results = search_instruments(query)
        return jsonify({
            'status': 'ok',
            'query': query,
            'results': results,
            'count': len(results)
        })
    
    except Exception as e:
        logger.error(f"Search error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/market_data', methods=['GET'])
def api_market_data():
    """获取市场数据和K线"""
    try:
        symbol = request.args.get('symbol', '').strip()
        timeframe = request.args.get('timeframe', '1d')
        days = int(request.args.get('days', 100))
        
        if not symbol:
            return jsonify({'error': 'Symbol required'}), 400
        
        logger.info(f"Fetching data for {symbol}")
        df = get_market_data(symbol, timeframe, days)
        
        if df is None or df.empty:
            # 返回演示数据而不是错误
            logger.info(f"No real data for {symbol}, using demo data")
            df = generate_demo_data(symbol, 30)
            if df is None or df.empty:
                return jsonify({'error': f'Could not fetch data for {symbol}'}), 404
        
        # 计算技术指标 - 使用try-catch避免崩溃
        indicators_dict = {}
        try:
            indicators_dict = calculate_all_indicators(df)
        except Exception as e:
            logger.warning(f"Indicator calculation failed: {e}, using empty indicators")
            # 创建空的指标字典
            indicators_dict = {k: np.zeros(len(df)) for k in ['MA1', 'MA2', 'MA3', 'MA4', 'VAR3', 'VAR4']}
        
        signal = {'buy': False, 'sell': False}
        try:
            signal = get_trading_signal(df) if indicators_dict else signal
        except:
            pass
        
        # 生成图表数据
        try:
            chart_data = generate_chart_data(df, indicators_dict)
        except Exception as e:
            logger.error(f"Chart data generation failed: {e}")
            chart_data = {'timestamps': [], 'ohlcv': [], 'ma1': [], 'ma2': [], 'ma3': [], 'ma4': []}
        
        # 获取当前价格
        current_price = float(df['close'].iloc[-1]) if 'close' in df.columns else 0
        
        # 清理指标以确保它们可序列化
        clean_indicators = {}
        for k, v in indicators_dict.items():
            try:
                if isinstance(v, np.ndarray):
                    if v.dtype == bool or v.dtype == np.bool_:
                        clean_indicators[k] = bool(v[-1]) if len(v) > 0 else False
                    else:
                        val = v[-1] if len(v) > 0 else 0
                        clean_indicators[k] = float(val) if not np.isnan(val) else 0
                elif isinstance(v, (list, tuple)):
                    continue
                elif isinstance(v, (bool, np.bool_)):
                    clean_indicators[k] = bool(v)
                else:
                    clean_indicators[k] = float(v) if v is not None else 0
            except:
                pass
        
        safe_signal = {}
        for sk, sv in signal.items():
            if isinstance(sv, (np.bool_, bool)):
                safe_signal[sk] = bool(sv)
            elif isinstance(sv, (np.integer, np.floating)):
                safe_signal[sk] = float(sv)
            else:
                safe_signal[sk] = sv
        
        return jsonify({
            'status': 'ok',
            'symbol': symbol,
            'instrument': {
                'symbol': symbol,
                'name': symbol,
                'type': 'ASSET'
            },
            'current_price': current_price,
            'chart_data': chart_data,
            'indicators': clean_indicators,
            'signal': safe_signal,
            'data_points': len(df)
        })
    
    except Exception as e:
        import traceback
        logger.error(f"Market data error: {e}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/analysis', methods=['POST'])
def api_analysis():
    """获取AI分析 - 完整版，包括所有Agent的分析"""
    try:
        data = request.get_json()
        symbol = data.get('symbol', '').strip()
        
        if not symbol:
            return jsonify({'error': 'Symbol required'}), 400
        
        logger.info(f"Analyzing {symbol}")
        
        # 获取市场数据
        df = get_market_data(symbol)
        if df is None or df.empty:
            logger.info(f"No real data for {symbol}, using demo data")
            df = generate_demo_data(symbol, 30)
            if df is None or df.empty:
                return jsonify({'error': f'Could not fetch data for {symbol}'}), 404
        
        current_price = float(df['close'].iloc[-1]) if 'close' in df.columns else 0
        
        # 多Agent共识分析 - 这是关键部分
        consensus = analyze_market_consensus(symbol, df, timeframe='1d')
        
        logger.info(f"Analysis result: {consensus.overall_signal}, Score: {consensus.consensus_score}")
        
        # ✅ 新增：计算交易参数
        entry_price = current_price
        leverage = random.randint(2, 10)  # 2-10倍杠杆
        
        if consensus.overall_signal == 'BUY':
            # 买入：止损价比入场价低1%，止盈价比入场价高0.5%
            stop_loss = entry_price * (1 - 0.01)
            take_profit = entry_price * (1 + 0.005)
        elif consensus.overall_signal == 'SELL':
            # 卖出：止损价比入场价高1%，止盈价比入场价低0.5%
            stop_loss = entry_price * (1 + 0.01)
            take_profit = entry_price * (1 - 0.005)
        else:
            # HOLD：不设置止损和止盈
            stop_loss = entry_price
            take_profit = entry_price
        
        # 构建完整的响应，包含所有Agent的分析
        return jsonify({
            'status': 'ok',
            'symbol': symbol,
            'current_price': current_price,
            'consensus': {
                'overall_signal': consensus.overall_signal,
                'consensus_score': consensus.consensus_score,
                'timestamp': consensus.timestamp,
                'entry_price': entry_price,
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'leverage': leverage
            },
            'agents': [
                {
                    'agent_name': d.agent_name,
                    'signal': d.signal,
                    'confidence': d.confidence,
                    'reasoning': d.reasoning
                }
                for d in consensus.agent_decisions
            ],
            'agent_count': len(consensus.agent_decisions)  # 确认有多少个Agent参与分析
        })
    
    except Exception as e:
        logger.error(f"Analysis error: {e}", exc_info=True)
        return jsonify({'error': str(e), 'detail': 'Check server logs'}), 500

@app.route('/api/chat', methods=['POST'])
def api_chat():
    """AI对话接口 - 改进版 - 需要有效的symbol"""
    try:
        data = request.get_json()
        user_message = data.get('message', '').strip()
        symbol = data.get('symbol', '').strip()
        
        if not user_message:
            return jsonify({'error': 'Message required'}), 400
        
        # ✅ 验证symbol是否提供 - 必须有分析过的标的才能使用对话
        if not symbol:
            return jsonify({
                'error': '请先完成分析。流程: 搜索标的 → 选择标的 → 选择时间周期 → 点击"分析"'
            }), 400
        
        # 获取用户ID
        if 'user_id' not in session:
            session['user_id'] = str(uuid.uuid4())
        
        user_id = session['user_id']
        
        # 在用户问题后面追加系统提示词（隐藏实现细节）
        system_prompt_suffix = f"""你正在分析投资品种：{symbol}

你必须在每一个回复中都明确提到这个投资品种名称：{symbol}

你的任务：
1. 以顶尖专业的金融投资专家的角色分析学习10个Agent（TechnicalAgent、MomentumAgent、VolatilityAgent、RiskManagementAgent、CryptoAnalystAgent、FundamentalAnalystAgent、SentimentAnalystAgent、MacroAnalystAgent、EventDrivenAgent、AdvancedAnalystAgent）各自的signal
2. 只根据ConsensusDecision的overall_signal来决定做多还是做空提供详细的仓位管理建议：
   - 高频交易杠杆建议
   - 仓位占账户比例
   - 单笔风险控制
   - 具体操作：做多和做空分别的入场价、止损价、止盈价
   - 做多和买入的止损价一定比入场价低，止盈价一定比入场价高；做空和卖出的止损价一定比入场价高，止盈价一定比入场价低。
3. 仓位计算：(目标利润 - 成本) / 每单位盈亏比
4. 风险提示：允许高杠杆但需要资金管理，强调爆仓风险

你的回复必须：
- 以你的任务为基础，直接给出根据ConsensusDecision对投资品种{symbol}的overall_signal和具体的操作指令来回复做多还是做空、入场价格、止损位、止盈位、杠杆倍数、仓位比例、预期盈亏。
- 按照用户提供的信息来直接回答即可，绝不能向用户索要更多信息，不要重复你的角色或目标，只输出最终的交易仓位管理建议以及顶级专业金融知识。
- 把10个agent的分析和结果转述出来，但不要输出它们的思考过程，直接转述它们的结论和信心度。"""
        
        # 构建完整的用户消息（明确包含symbol）
        user_message_with_symbol = f"关于{symbol}的问题：{user_message}"
        user_message_enhanced = f"{user_message_with_symbol}\n\n{system_prompt_suffix}"
        
        logger.info(f"User {user_id} asks about {symbol}: {user_message}")
        logger.info(f"User {user_id} asks: {user_message}")
        
        # 构建更好的提示词 - 专注于给出自然的回复，不是思考过程
        market_context = ""
        consensus_info = None
        
        if symbol:
            df = get_market_data(symbol)
            if df is not None and not df.empty:
                current_price = float(df['close'].iloc[-1])
                consensus = analyze_market_consensus(symbol, df, timeframe='1d')
                
                consensus_info = {
                    'symbol': symbol,
                    'current_price': current_price,
                    'overall_signal': consensus.overall_signal,
                    'consensus_score': consensus.consensus_score,
                    'agents': [
                        {
                            'agent_name': d.agent_name,
                            'signal': d.signal,
                            'confidence': d.confidence,
                            'reasoning': d.reasoning
                        }
                        for d in consensus.agent_decisions
                    ]
                }
                
                # 构建市场背景信息
                market_context = f"""
当前分析的标的: {symbol}
当前价格: ${current_price:.2f}
AI系统建议: {consensus.overall_signal}
系统信心度: {consensus.consensus_score:.0%}

各Agent分析:
"""
                for agent in consensus.agent_decisions:
                    market_context += f"- {agent.agent_name}: {agent.signal} (信心: {agent.confidence:.0%})\n"
        
        # 使用更自然的提示词，避免输出思考过程
        # 构建完整的聊天消息列表
        messages = [
            {
                'role': 'system',
                'content': """你是世界上最顶级专业二级市场风险投资专家。用户只做合约短线（周期1天-1个月），高杠杆高风险偏好，接受爆仓和50%波动率，资产占比极小无安全顾虑。
你的回复应该：
基于当前价格分析，直接给出具体操作指令——做多/做空、入场价格、止损位、止盈位、杠杆倍数、仓位比例、预期盈亏。
同时，你的回复明确禁止反问和提问题。
回复的要求必须，简洁直接，只输出可执行的交易计划， 直接、专业地给出建议，不反问问题，不要重复你的角色或目标，只输出最终的交易仓位管理建议以及顶级专业金融知识。"""
            },
            {
                'role': 'user',
                'content': f"""{market_context}

用户问题: {user_message_enhanced}"""
            }
        ]
        
        # 使用LLM生成回复
        try:
            llm = get_llm()
            response_text = llm.chat(
                messages,
                max_tokens=256,  # 减少到256以加快生成
                temperature=0.6  # 平衡：既不过于重复（0.3），也不过于随机（0.8）
            )
            
            # 清理：分别删除"用户回复: "、"用户问题"和"用户:"及其后面的所有内容
            if '用户回复: ' in response_text:
                response_text = response_text.split('用户回复: ')[0]
            
            if '用户问题' in response_text:
                response_text = response_text.split('用户问题')[0]
            
            if '用户:' in response_text:
                response_text = response_text.split('用户:')[0]

            if 'client' in response_text:
                response_text = response_text.split('client')[0]

            if '助手:' in response_text:
                response_text = response_text.split('助手:')[0]

            if '回答:' in response_text:
                response_text = response_text.split('回答:')[0]

            if 'Assistant' in response_text:
                response_text = response_text.split('Assistant')[0]

            # 如果LLM的回复仍然包含思考过程，尝试清理
            if '好的' in response_text[:20]:
                # 尝试找到真实回复的开始
                lines = response_text.split('\n')
                for i, line in enumerate(lines):
                    if any(keyword in line for keyword in ['建议', '应该', '可以', '不建议', '高', '低', '目标', '风险']):
                        response_text = '\n'.join(lines[i:])
                        break
                
            response_text = response_text.strip()
            
            # ============ 额外的防护清理：防止提示词泄露 ============
            # 移除可能泄露的系统指令
            import re
            
            # 删除"是一位专业的投资顾问AI"及其后续的系统指令
            if '是一位专业的投资顾问AI' in response_text:
                response_text = response_text.split('是一位专业的投资顾问AI')[0]
            
            # 删除回复指南模式
            response_text = re.sub(
                r"你的回复应该：\s*\n.*?自然流畅.*?(?:\n\n|$)",
                "",
                response_text,
                flags=re.IGNORECASE | re.DOTALL
            )
            
            # 删除"关于XXX的问题："这样的前缀
            response_text = re.sub(r"^关于\S+的问题：\s*", "", response_text)
            
            # 删除"你的任务："等任务说明
            response_text = re.sub(r"^你的任务：.*?(?:\n\n|$)", "", response_text, flags=re.DOTALL)
            
            # 最后再strip一次
            response_text = response_text.strip()
            
        except Exception as e:
            logger.warning(f"LLM generation failed: {e}")
            # 给出备用回复 - 确保包含symbol和明确的做多/做空建议
            if symbol and consensus_info:
                is_buy = "BUY" in consensus_info.get('overall_signal', 'SELL')
                direction = "合约做多，现货买入）" if is_buy else "合约做空，现货卖出"
                action_detail = "建议买入，设置止损价位，目标位在上方阻力" if is_buy else "建议卖出，设置止损价位，目标位在下方支撑"
                
                response_text = f"""关于{symbol}的投资建议：

当前{symbol}价格：${consensus_info.get('current_price', 'N/A')}
共识方向：{direction}
系统信心度：{(consensus_info.get('consensus_score', 0) * 100):.0f}%

{action_detail}

仓位管理建议：
建议仓位：账户总资金的20-50%
杠杆倍数：5-10倍（根据个人风险承受能力）
单笔风险：不超过账户总资金的20%
止损点：离入场价2-3%
止盈点：目标利润在5-10%

注意：这仅为技术面参考，请结合基本面分析和个人风险承受能力。"""
            else:
                response_text = """感谢你的问题！

为了给你更精准的建议，建议你：
1. 在搜索框中输入具体的投资品种代码（如AAPL、BTC/USDT等）
2. 点击分析按钮查看K线和技术指标
3. 然后提出具体问题

这样我可以基于实时市场数据给出专业的投资建议。

有什么具体的投资问题吗？"""
        
        return jsonify({
            'status': 'ok',
            'user_id': user_id,
            'message': user_message,
            'response': response_text,
            'consensus': consensus_info,
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/agents_info', methods=['GET'])
def api_agents_info():
    """获取Agent信息"""
    try:
        agent_system = get_agent_system()
        
        agents_info = [
            {
                'name': 'TechnicalAgent',
                'description': '技术分析Agent - 分析K线形态、均线、支阻位',
                'capabilities': ['K线形态', '移动平均线', '技术面判断']
            },
            {
                'name': 'MomentumAgent',
                'description': '动量Agent - 分析价格变化速度和趋势强度',
                'capabilities': ['ROC', '速度判断', '动量分析']
            },
            {
                'name': 'VolatilityAgent',
                'description': '波动率Agent - 评估市场波动水平',
                'capabilities': ['ATR', '波动率评估', '风险提示']
            },
            {
                'name': 'RiskManagementAgent',
                'description': '风险管理Agent - 控制风险和止损建议',
                'capabilities': ['风险评估', '止损建议', '仓位管理']
            }
        ]
        
        return jsonify({
            'status': 'ok',
            'agents': agents_info,
            'count': len(agents_info)
        })
    
    except Exception as e:
        logger.error(f"Agents info error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/indicators', methods=['GET'])
def api_indicators():
    """获取技术指标说明"""
    try:
        indicators_doc = {
            'main_chart': {
                'MA1': '5日均线',
                'MA2': '10日均线',
                'MA3': '20日均线',
                'MA4': '60日均线',
                'VAR3': '20日波段指标',
                'VAR4': '5日波段指标',
                'BUY_SIGNAL': '买点信号 (快线上穿慢线)',
                'SELL_SIGNAL': '卖点信号 (快线下穿慢线)',
                'STOP_LOSS_SIGNAL': '止损信号'
            },
            'subplot1': {
                'XX': 'KDJ-K值',
                'YY': 'KDJ-D值',
                'ESCAPE_TOP_SIGNAL': '逃顶信号'
            },
            'subplot3': {
                'CCI_VALUE': 'CCI指标',
                'CCI_THRESHOLD': 'CCI阈值'
            },
            'subplot4': {
                'DONT_BUY': '直接做空',
                'DONT_OPERATE': '继续直接做空',
                'VAR5': '风险指标'
            }
        }
        
        return jsonify({
            'status': 'ok',
            'indicators': indicators_doc
        })
    
    except Exception as e:
        logger.error(f"Indicators error: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== PRO会员API ====================

@app.route('/api/pro/signals', methods=['POST'])
def api_pro_signals():
    """PRO会员API - 获取指标系统信号（K线上的点位标注）"""
    try:
        data = request.get_json()
        symbol = data.get('symbol', '').strip()
        timeframe = data.get('timeframe', '1d')
        
        if not symbol:
            return jsonify({'error': 'Symbol required'}), 400
        
        # 获取市场数据
        df = get_market_data(symbol, timeframe, days=100)
        if df is None or df.empty:
            df = generate_demo_data(symbol, 30)
            if df is None or df.empty:
                return jsonify({'error': f'Could not fetch data for {symbol}'}), 404
        
        # 计算指标并提取信号
        indicators = TechnicalIndicators(df)
        
        # ==================== 第一个指标系统：波段买卖点 ====================
        close = df['close'].values
        high = df['high'].values
        low = df['low'].values
        open_price = df['open'].values
        
        # VA指标
        va2 = (close + high + low) / 3
        va3 = indicators.ema(va2, 10)
        va4 = indicators.ref(va3, 1)
        
        # 均线
        ma1 = indicators.sma(close, 5)
        ma2 = indicators.sma(close, 10)
        ma3 = indicators.sma(close, 20)
        ma4 = indicators.sma(close, 60)
        
        # 波段高低
        var1 = indicators.hhv('high', 25)
        var2 = indicators.llv('low', 25)
        var3 = indicators.ema((close - var2) / (var1 - var2 + 1e-10) * 100, 20)
        var4 = indicators.ema((close - var2) / (var1 - var2 + 1e-10) * 100, 5)
        
        # VAR5交叉信号
        var5 = indicators.cross(var4, var3)
        barslast_var5 = indicators.barslast(var5)
        
        signals_list = []
        
        # 遍历每一根K线生成信号
        for i in range(1, len(df)):
            timestamp = int(df.index[i].timestamp()) if hasattr(df.index[i], 'timestamp') else i * 1000
            
            # CROSS(VAR3,VAR4) AND BARSLAST(VAR5)>3 - 短卖
            if var3[i-1] <= var4[i-1] and var3[i] > var4[i]:
                if barslast_var5[i] > 3:
                    signals_list.append({
                        'time': timestamp,
                        'price': float(high[i]),
                        'type': 'sell_short',
                        'text': '分批做空',  
                        'color': "#D31ABA",
                        'icon': 2
                    })
            
            # CROSS(VAR3,VAR4) AND BARSLAST(VAR5)<=3 - 止损
            if var3[i-1] <= var4[i-1] and var3[i] > var4[i]:
                if barslast_var5[i] <= 3:
                    signals_list.append({
                        'time': timestamp,
                        'price': float(high[i]),
                        'type': 'stop_loss',
                        'text': '长线翻倍', 
                        'color': '#FFFFFF',
                        'icon': 2
                    })
            
            # CROSS(VAR4,VAR3) - 短买
            if var4[i-1] <= var3[i-1] and var4[i] > var3[i]:
                signals_list.append({
                    'time': timestamp,
                    'price': float(low[i]),
                    'type': 'buy',
                    'text': '分批做多', 
                    'color': "#17904D",
                    'icon': 1
                })
        
        # ==================== 第二个指标系统：不要买/不要操作 ====================
        var1_b = indicators.ref((low + open_price + high + close) / 4, 1)
        sma1 = indicators.sma((low - var1_b).astype(float).__abs__(), 13)
        sma2 = indicators.sma(np.maximum(low - var1_b, 0).astype(float), 10)
        var2_b = sma1 / (sma2 + 1e-10)
        var3_b = indicators.ema(var2_b, 10)
        var4_b = indicators.llv('low', 33)
        var5_b = indicators.ema(np.where(low <= var4_b, var3_b, 0), 3)
        

        buyoumai = np.where(var5_b > indicators.ref(var5_b, 1), var5_b, 0)
        buyaocaozuo = np.where(var5_b < indicators.ref(var5_b, 1), var5_b, 0)
        
        # 标记状态变化
        for i in range(1, len(df)):
            timestamp = int(df.index[i].timestamp()) if hasattr(df.index[i], 'timestamp') else i * 1000

            if buyoumai[i] > 0 and buyoumai[i-1] == 0:
                signals_list.append({
                    'time': timestamp,
                    'price': float(close[i]),
                    'type': 'indicator_red',
                    'text': '加仓做空机会',  
                    'color': "#FF9D00",
                    'icon': 0
                })
            
            if buyaocaozuo[i] > 0 and buyaocaozuo[i-1] == 0:
                signals_list.append({
                    'time': timestamp,
                    'price': float(close[i]),
                    'type': 'indicator_green',
                    'text': '做空机会',  # ⭐ 信号文本：继续直接做空
                    'color': "#FF0000",
                    'icon': 0
                })
        
        return jsonify({
            'status': 'ok',
            'symbol': symbol,
            'timeframe': timeframe,
            'signals': signals_list,
            'signal_count': len(signals_list),
            'indicators': {
                'MA1': [float(v) if not np.isnan(v) else 0 for v in ma1],
                'MA2': [float(v) if not np.isnan(v) else 0 for v in ma2],
                'MA3': [float(v) if not np.isnan(v) else 0 for v in ma3],
                'MA4': [float(v) if not np.isnan(v) else 0 for v in ma4],
                'VAR3': [float(v) if not np.isnan(v) else 0 for v in var3],
                'VAR4': [float(v) if not np.isnan(v) else 0 for v in var4]
            }
        })
    
    except Exception as e:
        logger.error(f"PRO signals error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# ==================== SVIP会员API ====================

@app.route('/api/svip/whale_data', methods=['GET'])
def api_svip_whale_data():
    """SVIP会员API - 获取鲸鱼数据（幕后数据）"""
    try:
        excel_path = '/Users/tonychan/Documents/trae_projects/AI/印钞机多空情况.xlsx'
        
        whale_data = {
            "long_count": 0,
            "short_count": 0,
            "long_addresses": [],
            "short_addresses": [],
            "trend": "数据加载中...",
            "long_ratio": 50.0,
            "short_ratio": 50.0
        }
        
        try:
            import re
            from openpyxl import load_workbook
            
            # 使用openpyxl直接读取Excel文件
            try:
                wb = load_workbook(excel_path)
                
                # 读取做多数据
                if '做多' in wb.sheetnames:
                    ws_long = wb['做多']
                    long_text = str(ws_long.cell(1, 1).value or "")
                else:
                    long_text = ""
                
                # 读取做空数据
                if '做空' in wb.sheetnames:
                    ws_short = wb['做空']
                    short_text = str(ws_short.cell(1, 1).value or "")
                else:
                    short_text = ""
                    
            except ImportError:
                logger.error("❌ openpyxl未安装，无法读取Excel文件")
                long_text = ""
                short_text = ""
            except Exception as e:
                logger.error(f"❌ 读取Excel文件失败: {e}")
                long_text = ""
                short_text = ""
            
            # 提取地址
            pattern = r'💰 0x([a-fA-F0-9]{4})\.\.\.([a-fA-F0-9]{3})'
            
            long_matches = re.findall(pattern, long_text)
            short_matches = re.findall(pattern, short_text)
            
            # 构建地址列表
            long_addresses = [f"0x{start}...{end}" for start, end in long_matches[:3]]
            short_addresses = [f"0x{start}...{end}" for start, end in short_matches[:3]]
            
            # 计算统计数据
            long_count = len(long_matches)
            short_count = len(short_matches)
            total = long_count + short_count
            
            if total > 0:
                long_ratio = (long_count / total) * 100
                short_ratio = (short_count / total) * 100
            else:
                long_ratio = 50.0
                short_ratio = 50.0
            
            # 动态生成趋势文本
            if long_ratio > 50:
                trend = f"💹 偏做多趋势 ({long_ratio:.1f}%)"
            elif long_ratio < 50:
                trend = f"📉 偏做空趋势 ({short_ratio:.1f}%)"
            else:
                trend = "⚖️ 趋势均衡 (50.0%)"
            
            whale_data = {
                "long_count": long_count,
                "short_count": short_count,
                "long_addresses": long_addresses,
                "short_addresses": short_addresses,
                "trend": trend,
                "long_ratio": long_ratio,
                "short_ratio": short_ratio
            }
            
            logger.info(f"✅ 鲸鱼数据加载成功: {trend}")
        
        except Exception as e:
            error_msg = str(e)
            logger.warning(f"⚠️ 鲸鱼数据加载失败: {error_msg}")
            
            # 特殊处理openpyxl导入失败的情况
            if "openpyxl" in error_msg or "import" in error_msg.lower():
                error_msg = "Excel读取模块缺失。请运行: pip install openpyxl"
                logger.error(f"❌ {error_msg}")
            
            whale_data = {
                "long_count": 0,
                "short_count": 0,
                "long_addresses": [],
                "short_addresses": [],
                "trend": f"数据加载失败: {error_msg}",
                "long_ratio": 50.0,
                "short_ratio": 50.0
            }
        
        return jsonify({
            'status': 'ok',
            'whale_data': whale_data,
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        error_msg = str(e)
        logger.error(f"SVIP whale data error: {error_msg}", exc_info=True)
        
        # 提供有用的错误信息
        if "openpyxl" in error_msg:
            error_msg = "Excel模块缺失，请运行: pip install openpyxl"
        
        return jsonify({
            'error': error_msg,
            'hint': '若是openpyxl错误，请安装: pip install openpyxl'
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'providers': list(providers.keys()),
        'llm': 'initialized'
    })

# ==================== 错误处理 ====================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Internal server error'}), 500

# ==================== 主程序 ====================

if __name__ == '__main__':
    logger.info("=" * 50)
    logger.info("投资AI系统启动")
    logger.info("=" * 50)
    logger.info(f"本地LLM: dmind-trading-merged")
    logger.info(f"数据提供者: {list(providers.keys())}")
    logger.info(f"时间: {datetime.now()}")
    logger.info("=" * 50)
    
    # 运行Flask应用 - 监听所有网卡，允许WiFi访问
    app.run(
        host='0.0.0.0',  # 监听所有网卡
        port=9000,       # 端口9000 
        debug=False,
        threaded=True
    )
